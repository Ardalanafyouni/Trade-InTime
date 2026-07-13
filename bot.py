import os
import asyncio
import logging
from io import BytesIO
from datetime import datetime, time as dtime

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler, filters, ContextTypes,
    JobQueue
)

from analyzer import CryptoAnalyzer
from chart_generator import generate_chart
from terms import TERMS, TERM_ALIASES, ALL_TERMS_FA, ALL_TERMS_EN, ALL_TERMS_RU
from journal import add_trade, get_trades, delete_trade, get_stats, load_journals, save_journals
from watchlist import generate_watchlist
from news import generate_news_message, get_fresh_breaking_news, fetch_news
from new_coins import generate_new_coins_message
from airdrops import generate_airdrops_message
from ai_analysis import generate_ai_analysis
from users import track_user, generate_users_report

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

WAITING_SYMBOL, WAITING_TIMEFRAME = range(2)
(J_SYMBOL, J_DIRECTION, J_ENTRY, J_SL, J_TP, J_SIZE, J_NOTE,
 J_CLOSE_ID, J_CLOSE_EXIT, J_CLOSE_PNL, J_DELETE_ID) = range(10, 21)

TEXTS = {
    'fa': {
        'welcome': "👋 *سلام! ربات تحلیل و ژورنال کریپتو*\n\n/analyze - تحلیل + چارت\n/watchlist - واچلیست هفتگی\n/journal - ژورنال معاملات\n/news - اخبار روز\n/newcoins - کوین‌های نوظهور\n/airdrops - ایردراپ‌های معتبر\n/source - منبع دیتای چارت\n/terms - اصطلاحات\n/lang - زبان",
        'enter_symbol': "🪙 نام ارز را وارد کنید:\nمثال: `BTC`, `ETH`, `SOL`",
        'choose_tf': "تایم‌فریم را انتخاب کنید:",
        'analyzing': "⏳ در حال تحلیل و رسم چارت",
        'error': "❌ خطا:", 'cancel': "❌ لغو شد.",
        'term_not_found': "❓ اصطلاح پیدا نشد. /terms را بزنید.",
        'choose_lang': "🌐 زبان را انتخاب کنید:",
        'lang_set': "✅ زبان فارسی انتخاب شد.",
        'watchlist_loading': "⏳ در حال دریافت داده‌های بازار...",
        'ai_button': "🤖 تحلیل هوش مصنوعی (اختصاصی)",
        'ai_loading': "⏳ در حال ترکیب دیتای تکنیکال با اخبار بروز و تولید تحلیل...",
        'ai_limit': "⛔️ شما به سقف {limit} تحلیل هوش مصنوعی در روز رسیده‌اید. فردا دوباره امتحان کنید.",
        'source_title': "📡 *منبع دیتای چارت و تحلیل*\n\nمنبع فعلی: *{current}*\n\nمی‌تونید انتخاب کنید کدوم منبع اول امتحان بشه. اگه اون منبع کوین رو نداشت، بقیه منابع خودکار امتحان می‌شن.",
        'source_auto': "🔄 خودکار (پیش‌فرض)",
        'source_changed': "✅ منبع دیتا روی *{source}* تنظیم شد.",
        'source_cmc_no_key': " (نیاز به کلید پولی CMC — فعلاً غیرفعاله)",
        'timeframes': {"1m": "1 دقیقه", "5m": "5 دقیقه", "15m": "15 دقیقه",
                       "1h": "1 ساعت", "4h": "4 ساعت", "1d": "روزانه", "1w": "هفتگی"},
        'journal': {
            'symbol_prompt': "🪙 نماد ارز:\nمثال: `BTC`",
            'no_trades': "📋 هیچ معامله‌ای ثبت نشده.",
            'trades_list_header': "📋 *آخرین معاملات:*\n\n",
            'no_stats': "📊 آماری موجود نیست.",
            'stats_template': "📊 *آمار:*\n\n• کل: {total}\n• بسته: {closed}\n• ✅ برنده: {win}\n• ❌ بازنده: {loss}\n• 🎯 وین ریت: {winrate}%\n• 💰 P&L کل: {total_pnl}$",
            'excel_caption': "📥 فایل اکسل ژورنال",
            'no_trades_delete': "هیچ معامله‌ای وجود ندارد.",
            'delete_header': "🗑 *کدام معامله را حذف کنید؟*\n\n",
            'enter_trade_number': "\nشماره معامله را وارد کنید:",
            'no_open_trades': "هیچ معامله بازی وجود ندارد.",
            'close_header': "شماره معامله برای بستن:\n\n",
            'direction_prompt': "جهت:",
            'entry_prompt': "💰 قیمت ورود:",
            'sl_prompt': "🛑 حد ضرر (SL):",
            'tp_prompt': "🎯 تارگت (TP):",
            'size_prompt': "📦 حجم ($):",
            'note_prompt': "📝 یادداشت (یا /skip):",
            'trade_saved': "✅ *معامله #{id} ثبت شد!*\n\n• {symbol} {direction}\n• ورود: {entry}\n• SL: {sl} | TP: {tp}",
            'trade_not_found': "❌ معامله #{id} پیدا نشد.",
            'trade_deleted': "✅ معامله #{id} حذف شد.",
            'invalid_number': "❌ شماره نامعتبر.",
            'exit_price_prompt': "💰 قیمت خروج:",
            'invalid': "❌ نامعتبر.",
            'pnl_prompt': "📊 P&L به دلار:\nمثال: `+150` یا `-80`",
            'profit': "✅ سود", 'loss': "❌ ضرر",
            'trade_closed': "{emoji}: {pnl}$\nمعامله #{id} بسته شد.",
            'invalid_value': "❌ مقدار نامعتبر.",
        },
    },
    'en': {
        'welcome': "👋 *Crypto Analysis & Journal Bot*\n\n/analyze - Analysis + Chart\n/watchlist - Weekly Watchlist\n/journal - Trade Journal\n/news - Crypto News\n/newcoins - New Coins\n/airdrops - Verified Airdrops\n/source - Chart Data Source\n/terms - Terms\n/lang - Language",
        'enter_symbol': "🪙 Enter coin symbol:\nExample: `BTC`, `ETH`, `SOL`",
        'choose_tf': "Select timeframe:",
        'analyzing': "⏳ Analyzing and generating chart",
        'error': "❌ Error:", 'cancel': "❌ Cancelled.",
        'term_not_found': "❓ Term not found. Use /terms",
        'choose_lang': "🌐 Choose your language:",
        'lang_set': "✅ English selected.",
        'watchlist_loading': "⏳ Fetching market data...",
        'ai_button': "🤖 AI Analysis (Custom)",
        'ai_loading': "⏳ Combining technical data with fresh news and generating analysis...",
        'ai_limit': "⛔️ You've reached the daily limit of {limit} AI analyses. Try again tomorrow.",
        'source_title': "📡 *Chart & Analysis Data Source*\n\nCurrent source: *{current}*\n\nChoose which source is tried first. If it doesn't have the coin, the rest are tried automatically.",
        'source_auto': "🔄 Auto (default)",
        'source_changed': "✅ Data source set to *{source}*.",
        'source_cmc_no_key': " (needs a paid CMC key — currently inactive)",
        'timeframes': {"1m": "1 Min", "5m": "5 Min", "15m": "15 Min",
                       "1h": "1 Hour", "4h": "4 Hour", "1d": "Daily", "1w": "Weekly"},
        'journal': {
            'symbol_prompt': "🪙 Coin symbol:\nExample: `BTC`",
            'no_trades': "📋 No trades recorded yet.",
            'trades_list_header': "📋 *Recent trades:*\n\n",
            'no_stats': "📊 No stats available.",
            'stats_template': "📊 *Stats:*\n\n• Total: {total}\n• Closed: {closed}\n• ✅ Wins: {win}\n• ❌ Losses: {loss}\n• 🎯 Win rate: {winrate}%\n• 💰 Total P&L: {total_pnl}$",
            'excel_caption': "📥 Journal Excel file",
            'no_trades_delete': "No trades to delete.",
            'delete_header': "🗑 *Which trade do you want to delete?*\n\n",
            'enter_trade_number': "\nEnter the trade number:",
            'no_open_trades': "No open trades.",
            'close_header': "Trade number to close:\n\n",
            'direction_prompt': "Direction:",
            'entry_prompt': "💰 Entry price:",
            'sl_prompt': "🛑 Stop Loss (SL):",
            'tp_prompt': "🎯 Take Profit (TP):",
            'size_prompt': "📦 Size ($):",
            'note_prompt': "📝 Note (or /skip):",
            'trade_saved': "✅ *Trade #{id} saved!*\n\n• {symbol} {direction}\n• Entry: {entry}\n• SL: {sl} | TP: {tp}",
            'trade_not_found': "❌ Trade #{id} not found.",
            'trade_deleted': "✅ Trade #{id} deleted.",
            'invalid_number': "❌ Invalid number.",
            'exit_price_prompt': "💰 Exit price:",
            'invalid': "❌ Invalid.",
            'pnl_prompt': "📊 P&L in dollars:\nExample: `+150` or `-80`",
            'profit': "✅ Profit", 'loss': "❌ Loss",
            'trade_closed': "{emoji}: {pnl}$\nTrade #{id} closed.",
            'invalid_value': "❌ Invalid value.",
        },
    },
    'ru': {
        'welcome': "👋 *Бот анализа и журнала крипто*\n\n/analyze - Анализ + График\n/watchlist - Вотч-лист\n/journal - Журнал сделок\n/news - Новости\n/newcoins - Новые монеты\n/airdrops - Проверенные аирдропы\n/source - Источник данных\n/terms - Термины\n/lang - Язык",
        'enter_symbol': "🪙 Введите символ:\nПример: `BTC`, `ETH`, `SOL`",
        'choose_tf': "Выберите таймфрейм:",
        'analyzing': "⏳ Анализирую и строю график",
        'error': "❌ Ошибка:", 'cancel': "❌ Отменено.",
        'term_not_found': "❓ Термин не найден. /terms",
        'choose_lang': "🌐 Выберите язык:",
        'lang_set': "✅ Русский выбран.",
        'watchlist_loading': "⏳ Получаю данные рынка...",
        'ai_button': "🤖 ИИ-анализ (персональный)",
        'ai_loading': "⏳ Объединяю технические данные со свежими новостями и генерирую анализ...",
        'ai_limit': "⛔️ Вы достигли дневного лимита {limit} ИИ-анализов. Попробуйте завтра.",
        'source_title': "📡 *Источник данных для графика и анализа*\n\nТекущий источник: *{current}*\n\nВыберите, какой источник пробовать первым. Если там нет монеты, остальные проверяются автоматически.",
        'source_auto': "🔄 Авто (по умолчанию)",
        'source_changed': "✅ Источник данных установлен на *{source}*.",
        'source_cmc_no_key': " (нужен платный ключ CMC — сейчас неактивен)",
        'timeframes': {"1m": "1 Мин", "5m": "5 Мин", "15m": "15 Мин",
                       "1h": "1 Час", "4h": "4 Часа", "1d": "День", "1w": "Неделя"},
        'journal': {
            'symbol_prompt': "🪙 Символ монеты:\nПример: `BTC`",
            'no_trades': "📋 Сделок пока не зарегистрировано.",
            'trades_list_header': "📋 *Последние сделки:*\n\n",
            'no_stats': "📊 Статистика недоступна.",
            'stats_template': "📊 *Статистика:*\n\n• Всего: {total}\n• Закрыто: {closed}\n• ✅ Прибыльных: {win}\n• ❌ Убыточных: {loss}\n• 🎯 Винрейт: {winrate}%\n• 💰 Общий P&L: {total_pnl}$",
            'excel_caption': "📥 Excel-файл журнала",
            'no_trades_delete': "Нет сделок для удаления.",
            'delete_header': "🗑 *Какую сделку удалить?*\n\n",
            'enter_trade_number': "\nВведите номер сделки:",
            'no_open_trades': "Нет открытых сделок.",
            'close_header': "Номер сделки для закрытия:\n\n",
            'direction_prompt': "Направление:",
            'entry_prompt': "💰 Цена входа:",
            'sl_prompt': "🛑 Стоп-лосс (SL):",
            'tp_prompt': "🎯 Тейк-профит (TP):",
            'size_prompt': "📦 Объём ($):",
            'note_prompt': "📝 Заметка (или /skip):",
            'trade_saved': "✅ *Сделка #{id} сохранена!*\n\n• {symbol} {direction}\n• Вход: {entry}\n• SL: {sl} | TP: {tp}",
            'trade_not_found': "❌ Сделка #{id} не найдена.",
            'trade_deleted': "✅ Сделка #{id} удалена.",
            'invalid_number': "❌ Неверный номер.",
            'exit_price_prompt': "💰 Цена выхода:",
            'invalid': "❌ Неверно.",
            'pnl_prompt': "📊 P&L в долларах:\nПример: `+150` или `-80`",
            'profit': "✅ Прибыль", 'loss': "❌ Убыток",
            'trade_closed': "{emoji}: {pnl}$\nСделка #{id} закрыта.",
            'invalid_value': "❌ Неверное значение.",
        },
    }
}

analyzer = CryptoAnalyzer()
user_langs = {}
news_subscribers = set()
newcoins_subscribers = set()
seen_news_links = set()

# ── AI Analysis rate limiting (in-memory, resets on deploy) ──
AI_DAILY_LIMIT = 5
ai_usage = {}  # uid -> {'date': date, 'count': int}

def check_and_use_ai_quota(uid):
    today = datetime.utcnow().date()
    entry = ai_usage.get(uid)
    if not entry or entry['date'] != today:
        entry = {'date': today, 'count': 0}
        ai_usage[uid] = entry
    if entry['count'] >= AI_DAILY_LIMIT:
        return False
    entry['count'] += 1
    return True

# ── Admin access (for /stats) ──
ADMIN_IDS = {int(x) for x in os.environ.get("ADMIN_IDS", "").split(",") if x.strip().isdigit()}

# Saved data-source preference (in-memory, per user)
user_source_pref = {}  # uid -> 'kucoin' / 'mexc' / 'coingecko' / 'coinmarketcap' / None (auto)


def is_admin(uid):
    return uid in ADMIN_IDS


async def run_blocking(func, *args, timeout=25, **kwargs):
    """Runs a blocking (sync) call in a worker thread with a hard timeout,
    so a slow/unresponsive exchange raises a catchable error instead of
    hanging the whole bot forever with no feedback to the user."""
    return await asyncio.wait_for(asyncio.to_thread(func, *args, **kwargs), timeout=timeout)

def get_lang(uid): return user_langs.get(uid, 'fa')
def t(uid, key): return TEXTS[get_lang(uid)][key]


# ── Excel Export ──
def export_to_excel(uid, lang='fa'):
    trades = get_trades(uid)
    stats = get_stats(uid)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trade Journal"
    headers_map = {
        'fa': ['#', 'تاریخ', 'ارز', 'جهت', 'ورود', 'SL', 'TP', 'حجم', 'خروج', 'P&L ($)', 'وضعیت', 'یادداشت'],
        'en': ['#', 'Date', 'Symbol', 'Direction', 'Entry', 'SL', 'TP', 'Size', 'Exit', 'P&L ($)', 'Status', 'Note'],
        'ru': ['#', 'Дата', 'Символ', 'Направление', 'Вход', 'SL', 'TP', 'Объём', 'Выход', 'P&L ($)', 'Статус', 'Заметка'],
    }
    headers = headers_map.get(lang, headers_map['en'])
    col_widths = [5, 18, 10, 10, 12, 12, 12, 10, 12, 12, 10, 25]
    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = f"Trade Journal — {datetime.utcnow().strftime('%Y-%m-%d')}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2962FF")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    if stats:
        ws.merge_cells('A2:L2')
        sc = ws['A2']
        sc.value = f"Total: {stats['total']} | Closed: {stats['closed']} | Win: {stats['win']} | Loss: {stats['loss']} | Winrate: {stats['winrate']}% | P&L: {stats['total_pnl']}$"
        sc.font = Font(bold=True, size=10, color="FFFFFF")
        sc.fill = PatternFill("solid", fgColor="263238")
        sc.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 22
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = w
    thin = Side(style='thin', color='2A2E39')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx, trade in enumerate(trades, 4):
        pnl = trade.get('pnl')
        row_bg = "1B3A1B" if pnl and pnl > 0 else "3A1B1B" if pnl and pnl < 0 else "1E222D"
        direction = trade.get('direction', '')
        dir_color = "00E676" if 'LONG' in direction else "FF5252"
        values = [trade.get('id',''), trade.get('date',''), trade.get('symbol',''), direction,
                  trade.get('entry',''), trade.get('sl',''), trade.get('tp',''), trade.get('size',''),
                  trade.get('exit_price',''), trade.get('pnl',''), trade.get('status','open'), trade.get('note','')]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.font = Font(color="FFFFFF", size=9)
            if col == 4: cell.font = Font(color=dir_color, bold=True, size=9)
            if col == 10 and pnl is not None:
                cell.font = Font(color="00E676" if pnl > 0 else "FF5252", bold=True, size=9)
        ws.row_dimensions[row_idx].height = 18
    ws.freeze_panes = 'A4'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Watchlist ──
async def watchlist_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    lang = get_lang(uid)
    msg = await update.message.reply_text(t(uid, 'watchlist_loading'))
    try:
        result, symbols_list = await run_blocking(generate_watchlist, lang, timeout=340)
        await msg.delete()

        keyboard = []
        row = []
        for i, sym in enumerate(symbols_list):
            row.append(InlineKeyboardButton(f"📊 {sym}", callback_data=f"wl_{sym}"))
            if len(row) == 2:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)

        reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
        await update.message.reply_text(result, parse_mode="Markdown", reply_markup=reply_markup)
    except Exception as e:
        logger.error(f"Watchlist error: {e}")
        await msg.edit_text(f"❌ خطا: {str(e)}")


async def newcoins_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    newcoins_subscribers.add(uid)
    lang = get_lang(uid)
    loading_text = {"fa": "⏳ در حال بررسی کوین‌های جدید...", "en": "⏳ Scanning new coins...", "ru": "⏳ Сканирую новые монеты..."}
    msg = await update.message.reply_text(loading_text.get(lang, loading_text['en']))
    try:
        result, symbols_list = generate_new_coins_message(lang)
        await msg.delete()

        keyboard = []
        row = []
        for sym in symbols_list:
            row.append(InlineKeyboardButton(f"📊 {sym}", callback_data=f"wl_{sym}"))
            if len(row) == 2:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None

        await update.message.reply_text(result, parse_mode="Markdown", reply_markup=reply_markup, disable_web_page_preview=True)
    except Exception as e:
        logger.error(f"New coins error: {e}")
        await msg.edit_text(f"❌ {str(e)}")


async def send_daily_newcoins(context):
    for uid in newcoins_subscribers:
        try:
            lang = get_lang(uid)
            result, symbols_list = generate_new_coins_message(lang)
            keyboard = []
            row = []
            for sym in symbols_list:
                row.append(InlineKeyboardButton(f"📊 {sym}", callback_data=f"wl_{sym}"))
                if len(row) == 2:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)
            reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
            await context.bot.send_message(chat_id=uid, text=result, parse_mode="Markdown", reply_markup=reply_markup, disable_web_page_preview=True)
        except:
            pass


async def airdrops_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    lang = get_lang(uid)
    loading_text = {"fa": "⏳ در حال بررسی ایردراپ‌های معتبر...", "en": "⏳ Scanning verified airdrops...", "ru": "⏳ Сканирую проверенные аирдропы..."}
    msg = await update.message.reply_text(loading_text.get(lang, loading_text['en']))
    try:
        result = generate_airdrops_message(lang)
        await msg.delete()
        await update.message.reply_text(result, parse_mode="Markdown", disable_web_page_preview=True)
    except Exception as e:
        logger.error(f"Airdrops error: {e}")
        await msg.edit_text(f"❌ {str(e)}")


async def news_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    first_time = uid not in news_subscribers
    news_subscribers.add(uid)
    lang = get_lang(uid)
    loading_text = {"fa": "⏳ در حال دریافت اخبار...", "en": "⏳ Fetching news...", "ru": "⏳ Получаю новости..."}
    msg = await update.message.reply_text(loading_text.get(lang, loading_text['en']))
    try:
        result = generate_news_message(lang)
        if first_time:
            items = fetch_news('en', limit=15)
            for item in items:
                seen_news_links.add(item['link'])
        await msg.delete()
        await update.message.reply_text(result, parse_mode="Markdown", disable_web_page_preview=True)
    except Exception as e:
        logger.error(f"News error: {e}")
        await msg.edit_text(f"❌ {str(e)}")


async def send_hourly_news(context):
    for uid in news_subscribers:
        try:
            lang = get_lang(uid)
            result = generate_news_message(lang)
            await context.bot.send_message(chat_id=uid, text=result, parse_mode="Markdown", disable_web_page_preview=True)
        except:
            pass


async def check_breaking_news(context):
    """Runs every 5 minutes to check for breaking news and alert instantly"""
    global seen_news_links
    if not news_subscribers:
        return
    try:
        fresh_items_en = get_fresh_breaking_news('en', seen_news_links)
        if not fresh_items_en:
            return

        for uid in news_subscribers:
            lang = get_lang(uid)
            fresh_items = get_fresh_breaking_news(lang, seen_news_links)
            for item in fresh_items[:3]:
                try:
                    alert_label = {"fa": "🚨 خبر فوری", "en": "🚨 BREAKING NEWS", "ru": "🚨 СРОЧНЫЕ НОВОСТИ"}
                    text = f"{alert_label.get(lang, alert_label['en'])}\n\n*{item['title']}*\n\n📡 {item['source']}\n🔗 {item['link']}"
                    await context.bot.send_message(chat_id=uid, text=text, parse_mode="Markdown", disable_web_page_preview=True)
                except:
                    pass

        for item in fresh_items_en:
            seen_news_links.add(item['link'])

        if len(seen_news_links) > 200:
            seen_news_links = set(list(seen_news_links)[-100:])
    except Exception as e:
        logger.error(f"Breaking news check error: {e}")


async def watchlist_quick_analyze(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    symbol = query.data.replace("wl_", "")
    context.user_data['symbol'] = symbol

    tfs = t(uid, 'timeframes')
    keyboard = [
        [InlineKeyboardButton(tfs["1m"], callback_data="1m"), InlineKeyboardButton(tfs["5m"], callback_data="5m"), InlineKeyboardButton(tfs["15m"], callback_data="15m")],
        [InlineKeyboardButton(tfs["1h"], callback_data="1h"), InlineKeyboardButton(tfs["4h"], callback_data="4h")],
        [InlineKeyboardButton(tfs["1d"], callback_data="1d"), InlineKeyboardButton(tfs["1w"], callback_data="1w")],
    ]
    await query.message.reply_text(f"✅ {symbol}USDT\n\n{t(uid, 'choose_tf')}", reply_markup=InlineKeyboardMarkup(keyboard))
    return WAITING_TIMEFRAME


# ── Journal ──
async def journal_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    stats = get_stats(uid)
    keyboard = [
        [InlineKeyboardButton("➕ ثبت معامله", callback_data="j_new"),
         InlineKeyboardButton("📋 لیست", callback_data="j_list")],
        [InlineKeyboardButton("✅ بستن معامله", callback_data="j_close"),
         InlineKeyboardButton("📊 آمار", callback_data="j_stats")],
        [InlineKeyboardButton("🗑 حذف معامله", callback_data="j_delete"),
         InlineKeyboardButton("📥 اکسپورت اکسل", callback_data="j_export")],
    ]
    text = "📒 *ژورنال معاملات*\n\n"
    if stats:
        text += f"• کل: {stats['total']} | بسته: {stats['closed']}\n• وین ریت: {stats['winrate']}%\n• P&L: {stats['total_pnl']}$"
    else:
        text += "هنوز معامله‌ای ثبت نشده."
    await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")


async def journal_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    data = query.data
    J = t(uid, 'journal')

    if data == "j_new":
        await query.message.reply_text(J['symbol_prompt'], parse_mode="Markdown")
        return J_SYMBOL

    elif data == "j_list":
        trades = get_trades(uid)
        if not trades:
            await query.message.reply_text(J['no_trades'])
            return
        text = J['trades_list_header']
        for trade in trades[-10:]:
            pnl = trade.get('pnl')
            status = trade.get('status', 'open')
            emoji = "🟢" if status == 'closed' and pnl and pnl > 0 else "🔴" if status == 'closed' and pnl and pnl <= 0 else "🟡"
            pnl_text = f" | P&L: {pnl}$" if pnl is not None else ""
            text += f"{emoji} #{trade['id']} {trade.get('symbol')} {trade.get('direction')} @ {trade.get('entry')}{pnl_text}\n"
        await query.message.reply_text(text, parse_mode="Markdown")

    elif data == "j_stats":
        stats = get_stats(uid)
        if not stats:
            await query.message.reply_text(J['no_stats'])
            return
        text = J['stats_template'].format(**stats)
        await query.message.reply_text(text, parse_mode="Markdown")

    elif data == "j_export":
        lang = get_lang(uid)
        buf = export_to_excel(uid, lang)
        await query.message.reply_document(document=buf, filename=f"journal_{datetime.utcnow().strftime('%Y%m%d')}.xlsx", caption=J['excel_caption'])

    elif data == "j_delete":
        trades = get_trades(uid)
        if not trades:
            await query.message.reply_text(J['no_trades_delete'])
            return
        text = J['delete_header']
        for trade in trades[-10:]:
            pnl = trade.get('pnl')
            pnl_text = f" | P&L: {pnl}$" if pnl is not None else ""
            text += f"#{trade['id']} {trade.get('symbol')} {trade.get('direction')} @ {trade.get('entry')}{pnl_text}\n"
        text += J['enter_trade_number']
        await query.message.reply_text(text, parse_mode="Markdown")
        return J_DELETE_ID

    elif data == "j_close":
        trades = get_trades(uid)
        open_trades = [t for t in trades if t.get('status') == 'open']
        if not open_trades:
            await query.message.reply_text(J['no_open_trades'])
            return
        text = J['close_header']
        for t_ in open_trades[-5:]:
            text += f"#{t_['id']} {t_.get('symbol')} {t_.get('direction')} @ {t_.get('entry')}\n"
        await query.message.reply_text(text)
        return J_CLOSE_ID


# ── Journal conversation handlers ──
async def j_symbol(update, context):
    uid = update.effective_user.id
    context.user_data['j_symbol'] = update.message.text.strip().upper()
    keyboard = [[InlineKeyboardButton("📈 LONG", callback_data="jd_long"), InlineKeyboardButton("📉 SHORT", callback_data="jd_short")]]
    await update.message.reply_text(t(uid, 'journal')['direction_prompt'], reply_markup=InlineKeyboardMarkup(keyboard))
    return J_DIRECTION

async def j_direction(update, context):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    context.user_data['j_direction'] = "LONG" if query.data == "jd_long" else "SHORT"
    await query.message.reply_text(t(uid, 'journal')['entry_prompt'])
    return J_ENTRY

async def j_entry(update, context):
    uid = update.effective_user.id
    context.user_data['j_entry'] = update.message.text.strip()
    await update.message.reply_text(t(uid, 'journal')['sl_prompt'])
    return J_SL

async def j_sl(update, context):
    uid = update.effective_user.id
    context.user_data['j_sl'] = update.message.text.strip()
    await update.message.reply_text(t(uid, 'journal')['tp_prompt'])
    return J_TP

async def j_tp(update, context):
    uid = update.effective_user.id
    context.user_data['j_tp'] = update.message.text.strip()
    await update.message.reply_text(t(uid, 'journal')['size_prompt'])
    return J_SIZE

async def j_size(update, context):
    uid = update.effective_user.id
    context.user_data['j_size'] = update.message.text.strip()
    await update.message.reply_text(t(uid, 'journal')['note_prompt'])
    return J_NOTE

async def j_note(update, context):
    uid = update.effective_user.id
    J = t(uid, 'journal')
    note = "" if update.message.text.strip() == '/skip' else update.message.text.strip()
    trade = {
        'symbol': context.user_data.get('j_symbol'),
        'direction': context.user_data.get('j_direction'),
        'entry': context.user_data.get('j_entry'),
        'sl': context.user_data.get('j_sl'),
        'tp': context.user_data.get('j_tp'),
        'size': context.user_data.get('j_size'),
        'note': note, 'status': 'open', 'exit_price': None, 'pnl': None,
    }
    trade_id = add_trade(uid, trade)
    await update.message.reply_text(
        J['trade_saved'].format(id=trade_id, symbol=trade['symbol'], direction=trade['direction'],
                                 entry=trade['entry'], sl=trade['sl'], tp=trade['tp']),
        parse_mode="Markdown"
    )
    return ConversationHandler.END

async def j_delete_id(update, context):
    uid = update.effective_user.id
    J = t(uid, 'journal')
    try:
        trade_id = int(update.message.text.strip())
        trades = get_trades(uid)
        trade_ids = [t['id'] for t in trades]
        if trade_id not in trade_ids:
            await update.message.reply_text(J['trade_not_found'].format(id=trade_id))
            return ConversationHandler.END
        delete_trade(uid, trade_id)
        await update.message.reply_text(J['trade_deleted'].format(id=trade_id))
    except:
        await update.message.reply_text(J['invalid_number'])
    return ConversationHandler.END

async def j_close_id(update, context):
    uid = update.effective_user.id
    J = t(uid, 'journal')
    try:
        context.user_data['j_close_id'] = int(update.message.text.strip())
        await update.message.reply_text(J['exit_price_prompt'])
        return J_CLOSE_EXIT
    except:
        await update.message.reply_text(J['invalid'])
        return ConversationHandler.END

async def j_close_exit(update, context):
    uid = update.effective_user.id
    context.user_data['j_close_exit'] = update.message.text.strip()
    await update.message.reply_text(t(uid, 'journal')['pnl_prompt'], parse_mode="Markdown")
    return J_CLOSE_PNL

async def j_close_pnl(update, context):
    uid = update.effective_user.id
    J = t(uid, 'journal')
    try:
        pnl = float(update.message.text.strip().replace('+', ''))
        trade_id = context.user_data.get('j_close_id')
        data = load_journals()
        for trade in data.get(str(uid), []):
            if trade.get('id') == trade_id:
                trade['status'] = 'closed'
                trade['exit_price'] = context.user_data.get('j_close_exit')
                trade['pnl'] = pnl
                break
        save_journals(data)
        emoji = J['profit'] if pnl > 0 else J['loss']
        await update.message.reply_text(J['trade_closed'].format(emoji=emoji, pnl=pnl, id=trade_id))
    except:
        await update.message.reply_text(J['invalid_value'])
    return ConversationHandler.END


# ── Analyze handlers ──
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    user = update.effective_user
    track_user(uid, user.username, user.first_name, started=True)
    await update.message.reply_text(t(uid, 'welcome'), parse_mode="Markdown")

async def track_activity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Runs on every incoming update (message or button tap) in a low-priority
    group so it never blocks the normal handlers. Lets /stats show anyone who
    has used the bot at all, not just people who explicitly ran /start."""
    user = update.effective_user
    if user:
        track_user(user.id, user.username, user.first_name)


SOURCE_DISPLAY_NAMES = {'kucoin': 'KuCoin', 'mexc': 'MEXC', 'coingecko': 'CoinGecko', 'coinmarketcap': 'CoinMarketCap'}

async def source_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    lang = get_lang(uid)
    current_key = user_source_pref.get(uid)
    current_label = SOURCE_DISPLAY_NAMES.get(current_key, t(uid, 'source_auto')) if current_key else t(uid, 'source_auto')

    keyboard = [[InlineKeyboardButton(t(uid, 'source_auto'), callback_data="src_auto")]]
    row = []
    for key in ['kucoin', 'mexc', 'coingecko', 'coinmarketcap']:
        label = SOURCE_DISPLAY_NAMES[key]
        if key == 'coinmarketcap' and not analyzer.CMC_API_KEY:
            label += t(uid, 'source_cmc_no_key')
        row.append(InlineKeyboardButton(label, callback_data=f"src_{key}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)

    await update.message.reply_text(
        t(uid, 'source_title').format(current=current_label),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    choice = query.data.replace("src_", "")

    if choice == "auto":
        user_source_pref.pop(uid, None)
        label = t(uid, 'source_auto')
    else:
        user_source_pref[uid] = choice
        label = SOURCE_DISPLAY_NAMES.get(choice, choice)

    await query.edit_message_text(t(uid, 'source_changed').format(source=label), parse_mode="Markdown")


async def whoami_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    uid = user.id
    admin_note = "✅ شما تو لیست ADMIN_IDS هستید." if is_admin(uid) else "❌ شما ادمین نیستید (یا ADMIN_IDS هنوز درست تنظیم نشده)."
    await update.message.reply_text(
        f"🆔 آیدی عددی تلگرام شما: {uid}\n\n{admin_note}"
    )


async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not is_admin(uid):
        return
    lang = get_lang(uid)
    await update.message.reply_text(generate_users_report(lang))


async def lang_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    keyboard = [[InlineKeyboardButton("🇮🇷 فارسی", callback_data="lang_fa"), InlineKeyboardButton("🇬🇧 English", callback_data="lang_en"), InlineKeyboardButton("🇷🇺 Русский", callback_data="lang_ru")]]
    await update.message.reply_text(t(uid, 'choose_lang'), reply_markup=InlineKeyboardMarkup(keyboard))

async def set_lang(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    lang = query.data.split('_')[1]
    user_langs[uid] = lang
    await query.edit_message_text(TEXTS[lang]['lang_set'])

async def terms_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    lang = get_lang(uid)
    msg = ALL_TERMS_FA if lang == 'fa' else ALL_TERMS_EN if lang == 'en' else ALL_TERMS_RU
    await update.message.reply_text(msg, parse_mode="Markdown")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    lang = get_lang(uid)
    text = update.message.text.strip().lower()
    term_key = TERM_ALIASES.get(text) or TERM_ALIASES.get(text.upper())
    if not term_key and text.upper() in TERMS.get(lang, {}):
        term_key = text.upper()
    if term_key:
        lang_terms = TERMS.get(lang, TERMS['en'])
        explanation = lang_terms.get(term_key) or TERMS['en'].get(term_key, t(uid, 'term_not_found'))
        await update.message.reply_text(explanation, parse_mode="Markdown")
    else:
        await update.message.reply_text(t(uid, 'term_not_found'))

async def analyze_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    await update.message.reply_text(t(uid, 'enter_symbol'), parse_mode="Markdown")
    return WAITING_SYMBOL

async def receive_symbol(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    symbol = update.message.text.strip().upper()
    context.user_data['symbol'] = symbol
    tfs = t(uid, 'timeframes')
    keyboard = [
        [InlineKeyboardButton(tfs["1m"], callback_data="1m"), InlineKeyboardButton(tfs["5m"], callback_data="5m"), InlineKeyboardButton(tfs["15m"], callback_data="15m")],
        [InlineKeyboardButton(tfs["1h"], callback_data="1h"), InlineKeyboardButton(tfs["4h"], callback_data="4h")],
        [InlineKeyboardButton(tfs["1d"], callback_data="1d"), InlineKeyboardButton(tfs["1w"], callback_data="1w")],
    ]
    await update.message.reply_text(f"✅ {symbol}USDT\n\n{t(uid, 'choose_tf')}", reply_markup=InlineKeyboardMarkup(keyboard))
    return WAITING_TIMEFRAME

async def receive_timeframe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    timeframe = query.data
    symbol = context.user_data.get('symbol', 'BTC')
    lang = get_lang(uid)
    tfs = t(uid, 'timeframes')
    await query.edit_message_text(f"{t(uid, 'analyzing')} {symbol}USDT ...")
    pref_source = user_source_pref.get(uid)
    try:
        df = await run_blocking(analyzer.fetch_ohlcv, symbol, timeframe, limit=300, timeout=65, preferred_source=pref_source)
        patterns = analyzer.detect_patterns(df, lang)
        trend_label, trend_type = analyzer.determine_trend(df, lang)
        fib_levels, _, _ = analyzer.calc_fibonacci(df)
        supports, resistances = analyzer.find_support_resistance(df)
        scores = analyzer.compute_signal(df, patterns, trend_type)
        chart_buf = generate_chart(df, symbol, timeframe, patterns, trend_label, trend_type, fib_levels, supports, resistances, scores)
        text_result = await run_blocking(analyzer.analyze, symbol, timeframe, lang, preferred_source=pref_source)
        ai_keyboard = InlineKeyboardMarkup([[InlineKeyboardButton(t(uid, 'ai_button'), callback_data=f"ai_{symbol}_{timeframe}")]])
        caption = f"📊 {symbol}USDT | {tfs.get(timeframe)}"
        if analyzer.last_source and analyzer.last_source != 'kucoin':
            source_label = {"mexc": "MEXC", "coingecko": "CoinGecko", "coinmarketcap": "CoinMarketCap"}.get(analyzer.last_source, analyzer.last_source)
            src_note = {"fa": f"\n📡 منبع: {source_label} (روی KuCoin موجود نبود)",
                        "en": f"\n📡 Source: {source_label} (not available on KuCoin)",
                        "ru": f"\n📡 Источник: {source_label} (недоступно на KuCoin)"}
            caption += src_note.get(lang, src_note['en'])
        await query.message.reply_photo(photo=chart_buf, caption=caption)
        await query.message.reply_text(text_result, parse_mode="Markdown", reply_markup=ai_keyboard)
    except asyncio.TimeoutError:
        timeout_msg = {"fa": "⏱ صرافی خیلی دیر جواب داد. لطفاً چند لحظه دیگه دوباره امتحان کن.",
                       "en": "⏱ The exchange took too long to respond. Please try again in a moment.",
                       "ru": "⏱ Биржа слишком долго отвечала. Пожалуйста, попробуйте ещё раз через минуту."}
        await query.message.reply_text(timeout_msg.get(lang, timeout_msg['en']))
    except Exception as e:
        logger.error(f"Error: {e}")
        await query.message.reply_text(f"{t(uid, 'error')} {str(e)}")
    return ConversationHandler.END


async def ai_analysis_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    lang = get_lang(uid)

    try:
        _, symbol, timeframe = query.data.split('_', 2)
    except ValueError:
        return

    if not check_and_use_ai_quota(uid):
        await query.message.reply_text(t(uid, 'ai_limit').format(limit=AI_DAILY_LIMIT))
        return

    msg = await query.message.reply_text(t(uid, 'ai_loading'))
    try:
        df = await run_blocking(analyzer.fetch_ohlcv, symbol, timeframe, limit=300, timeout=65, preferred_source=user_source_pref.get(uid))
        patterns = analyzer.detect_patterns(df, lang)
        trend_label, trend_type = analyzer.determine_trend(df, lang)
        fib_levels, _, _ = analyzer.calc_fibonacci(df)
        supports, resistances = analyzer.find_support_resistance(df)
        scores = analyzer.compute_signal(df, patterns, trend_type)
        atr = analyzer.calc_atr(df).iloc[-1]
        price = df['close'].iloc[-1]

        result_text, err = generate_ai_analysis(
            symbol, timeframe, price, trend_label, scores, patterns,
            fib_levels, supports, resistances, atr, lang
        )
        await msg.delete()
        if err:
            await query.message.reply_text(f"❌ {err}")
        else:
            await query.message.reply_text(result_text, disable_web_page_preview=True)
    except Exception as e:
        logger.error(f"AI analysis error: {e}")
        await msg.edit_text(f"{t(uid, 'error')} {str(e)}")

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    await update.message.reply_text(t(uid, 'cancel'))
    return ConversationHandler.END


# ── Main ──
def main():
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise ValueError("TELEGRAM_BOT_TOKEN not set!")

    app = ApplicationBuilder().token(token).build()

    app.job_queue.run_daily(
        send_daily_newcoins,
        time=dtime(9, 0, 0),
        name="daily_newcoins"
    )

    app.job_queue.run_repeating(
        check_breaking_news,
        interval=300,
        first=90,
        name="breaking_news_check"
    )

    analyze_conv = ConversationHandler(
        entry_points=[
            CommandHandler("analyze", analyze_command),
            CallbackQueryHandler(watchlist_quick_analyze, pattern="^wl_"),
        ],
        states={
            WAITING_SYMBOL: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_symbol)],
            WAITING_TIMEFRAME: [CallbackQueryHandler(receive_timeframe, pattern="^(1m|5m|15m|1h|4h|1d|1w)$")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )

    journal_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(journal_callback, pattern="^j_")],
        states={
            J_SYMBOL: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_symbol)],
            J_DIRECTION: [CallbackQueryHandler(j_direction, pattern="^jd_")],
            J_ENTRY: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_entry)],
            J_SL: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_sl)],
            J_TP: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_tp)],
            J_SIZE: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_size)],
            J_NOTE: [MessageHandler(filters.TEXT, j_note)],
            J_DELETE_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_delete_id)],
            J_CLOSE_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_close_id)],
            J_CLOSE_EXIT: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_close_exit)],
            J_CLOSE_PNL: [MessageHandler(filters.TEXT & ~filters.COMMAND, j_close_pnl)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CommandHandler("whoami", whoami_command))
    app.add_handler(CommandHandler("source", source_command))
    app.add_handler(CommandHandler("lang", lang_command))
    app.add_handler(CommandHandler("terms", terms_command))
    app.add_handler(CommandHandler("journal", journal_command))
    app.add_handler(CommandHandler("watchlist", watchlist_command))
    app.add_handler(CommandHandler("news", news_command))
    app.add_handler(CommandHandler("newcoins", newcoins_command))
    app.add_handler(CommandHandler("airdrops", airdrops_command))
    app.add_handler(CallbackQueryHandler(set_lang, pattern="^lang_"))
    app.add_handler(CallbackQueryHandler(ai_analysis_callback, pattern="^ai_"))
    app.add_handler(CallbackQueryHandler(source_callback, pattern="^src_"))

    # Low-priority group: tracks any user activity (message or button tap)
    # for /stats, without blocking the normal handlers above.
    app.add_handler(MessageHandler(filters.ALL, track_activity), group=1)
    app.add_handler(CallbackQueryHandler(track_activity), group=1)
    app.add_handler(analyze_conv)
    app.add_handler(journal_conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    logger.info("Bot is running...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
# THIS IS JUST THE ADDITIONS - NOT A STANDALONE FILE


